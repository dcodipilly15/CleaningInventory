import http.server
import webbrowser
import urllib.parse
import threading
import socket
import json
import re
import os
import sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("\nMissing package. Please run:\n\n    pip3 install openpyxl\n")
    sys.exit(1)

SCRIPT_DIR     = os.path.dirname(os.path.abspath(__file__))
PORT           = 8765
ADMIN_PORT     = 8766
LOCATIONS_FILE = os.path.join(SCRIPT_DIR, "locations.json")
COMBINED_FILE  = os.path.join(SCRIPT_DIR, "Inventory.xlsx")

# Default supplies used when creating a new location
TEMPLATE_ITEMS = [
    ("All-Purpose Cleaner",  20,  "Bottles"),
    ("Glass Cleaner",        15,  "Bottles"),
    ("Disinfectant Spray",   18,  "Cans"),
    ("Bathroom Cleaner",     10,  "Bottles"),
    ("Floor Cleaner",         8,  "Gallons"),
    ("Bleach",                6,  "Gallons"),
    ("Dish Soap",            12,  "Bottles"),
    ("Hand Soap",            24,  "Bottles"),
    ("Mop Heads",            10,  "Units"),
    ("Microfiber Cloths",    50,  "Cloths"),
    ("Scrub Brushes",        15,  "Units"),
    ("Rubber Gloves",        30,  "Pairs"),
    ("Trash Bags (Large)",  100,  "Bags"),
    ("Trash Bags (Small)",  200,  "Bags"),
    ("Paper Towels",         40,  "Rolls"),
    ("Toilet Paper",         96,  "Rolls"),
]

# Original separate-file configs — used only for one-time migration
_OLD_CONFIGS = {
    "milmont": {
        "label":    "Milmont",
        "file":     os.path.join(SCRIPT_DIR, "Inventory for Milmont.xlsm"),
        "sheet":    "Cleaning Supplies",
        "qty_col":  2,
        "keep_vba": True,
    },
    "3k": {
        "label":    "3K",
        "file":     os.path.expanduser("~/Downloads/Inventory for 3K.xlsx"),
        "sheet":    "Sheet1",
        "qty_col":  7,
        "keep_vba": False,
    },
}

# Default config pointing to the combined file
_HARDCODED = {
    "milmont": {
        "label":    "Milmont",
        "file":     COMBINED_FILE,
        "sheet":    "Milmont",
        "qty_col":  2,
        "unit_col": 3,
        "keep_vba": False,
    },
    "3k": {
        "label":    "3K",
        "file":     COMBINED_FILE,
        "sheet":    "3K",
        "qty_col":  2,
        "unit_col": 3,
        "keep_vba": False,
    },
}


def _init_inventories():
    if os.path.exists(LOCATIONS_FILE):
        with open(LOCATIONS_FILE) as f:
            data = json.load(f)
        # Upgrade milmont/3k entries to combined file if it now exists
        if os.path.exists(COMBINED_FILE):
            for key in ("milmont", "3k"):
                if key in data and key in _HARDCODED:
                    data[key] = _HARDCODED[key]
        return data
    return dict(_HARDCODED)


INVENTORIES = _init_inventories()


def save_locations():
    with open(LOCATIONS_FILE, "w") as f:
        json.dump(INVENTORIES, f, indent=2)


def migrate_to_combined():
    if os.path.exists(COMBINED_FILE):
        return
    print("  Combining Excel files into one...")
    wb_new = openpyxl.Workbook()
    wb_new.remove(wb_new.active)
    for key, old_cfg in _OLD_CONFIGS.items():
        label = old_cfg["label"]
        ws = wb_new.create_sheet(label)
        ws.cell(1, 1).value = "Item Name"
        ws.cell(1, 2).value = "Quantity"
        ws.cell(1, 3).value = "Unit"
        if os.path.exists(old_cfg["file"]):
            try:
                items = load_inventory(old_cfg)
                for i, item in enumerate(items, start=2):
                    ws.cell(i, 1).value = item["name"]
                    ws.cell(i, 2).value = item["qty"]
                    ws.cell(i, 3).value = item["unit"] or ""
                print(f"  Migrated {label} data from existing file.")
            except Exception as e:
                print(f"  Could not read {label} ({e}), using template defaults.")
                for i, (name, qty, unit) in enumerate(TEMPLATE_ITEMS, start=2):
                    ws.cell(i, 1).value = name
                    ws.cell(i, 2).value = qty
                    ws.cell(i, 3).value = unit
        else:
            print(f"  No existing file for {label}, using template defaults.")
            for i, (name, qty, unit) in enumerate(TEMPLATE_ITEMS, start=2):
                ws.cell(i, 1).value = name
                ws.cell(i, 2).value = qty
                ws.cell(i, 3).value = unit
    wb_new.save(COMBINED_FILE)
    # Update INVENTORIES and save new locations.json
    for key in ("milmont", "3k"):
        if key in INVENTORIES and key in _HARDCODED:
            INVENTORIES[key] = _HARDCODED[key]
    save_locations()
    print(f"  Created: Inventory.xlsx")


def get_template_items():
    for cfg in INVENTORIES.values():
        try:
            items = load_inventory(cfg)
            if items:
                return [(item["name"], item["qty"], item["unit"]) for item in items]
        except Exception:
            continue
    return TEMPLATE_ITEMS


def create_location_sheet(label):
    """Add a new sheet to COMBINED_FILE with the same items as existing locations, qty all 0."""
    # Use the first working location as the master item list
    master = []
    for cfg in INVENTORIES.values():
        try:
            items = get_items(cfg)
            if items:
                master = [(item["name"], item["unit"]) for item in items]
                break
        except Exception:
            continue
    if not master:
        master = [(name, unit) for name, _qty, unit in TEMPLATE_ITEMS]

    wb = openpyxl.load_workbook(COMBINED_FILE)
    ws = wb.create_sheet(label)
    ws.cell(1, 1).value = None
    for i, (name, unit) in enumerate(master, start=2):
        ws.cell(i, 1).value = name
        ws.cell(i, 2).value = 0
        ws.cell(i, 3).value = unit
    wb.save(COMBINED_FILE)


def migrate_separate_locations():
    """Move any location stored in a separate file into COMBINED_FILE as a new sheet."""
    changed = False
    wb = None
    for key, cfg in list(INVENTORIES.items()):
        if cfg["file"] == COMBINED_FILE:
            continue
        label = cfg["label"]
        # Read existing data from the separate file
        try:
            items = load_inventory(cfg)
        except Exception:
            items = []
        # Open/reuse the combined workbook
        if wb is None:
            wb = openpyxl.load_workbook(COMBINED_FILE)
        if label not in wb.sheetnames:
            ws = wb.create_sheet(label)
            ws.cell(1, 1).value = None
            for i, item in enumerate(items, start=2):
                ws.cell(i, 1).value = item["name"]
                ws.cell(i, 2).value = item["qty"]
                ws.cell(i, 3).value = item["unit"] or ""
            print(f"  Migrated {label} into Inventory.xlsx")
        INVENTORIES[key] = {
            "label":    label,
            "file":     COMBINED_FILE,
            "sheet":    label,
            "qty_col":  2,
            "unit_col": 3,
            "keep_vba": False,
        }
        changed = True
    if wb is not None:
        wb.save(COMBINED_FILE)
    if changed:
        save_locations()


CATEGORIES_FILE = os.path.join(SCRIPT_DIR, "categories.json")

DEFAULT_CATEGORIES = {
    "Degreaser":                                    "Cleaning Sprays",
    "Multipurpose Cleaner Spray":                   "Cleaning Sprays",
    "Glass Cleaner":                                "Cleaning Sprays",
    "Stainless Steel Cleaner":                      "Cleaning Sprays",
    "Wood Furniture Cleaner":                       "Cleaning Sprays",
    "S.O.S Steel Pads":                             "Scrubbing & Sponges",
    "Scrub Mommy (Dual-Sided Scrubber + Sponge)":  "Scrubbing & Sponges",
    "Mr. Clean Magic Erasers":                      "Scrubbing & Sponges",
    "Lysol Cleaning Wipes":                         "Wipes",
    "Multipurpose Cleaning Wipes":                  "Wipes",
    "Dish Soap":                                    "Soaps & Detergents",
    "Dishwasher Detergent":                         "Soaps & Detergents",
    "Clorox ToiletWand Starter Kit":                "Toilet & Bathroom",
    "Toilet Wand Refills":                          "Toilet & Bathroom",
    "Lysol Toilet Cleaning Gel":                    "Toilet & Bathroom",
    "Swiffer Refills":                              "Mopping & Sweeping",
    "Swiffer WetJet":                               "Mopping & Sweeping",
    "Swiffer Mopping Kit":                          "Mopping & Sweeping",
    "Microfiber Towels":                            "Cloths & Towels",
    "Paper Towels":                                 "Cloths & Towels",
    "Small Trash Bags":                             "Bags",
    "Large Trash Bags":                             "Bags",
    "Available Target Bags":                        "Bags",
    "Nitrile Gloves (50-count)":                    "Gloves & PPE",
    "Gloves":                                       "Gloves & PPE",
    "Gardening Gloves":                             "Gloves & PPE",
    "Room Air Freshener":                           "Air Care",
    "Removable Vacuum Batteries":                   "Miscellaneous",
    "Foam Grip Tape":                               "Miscellaneous",
    "Post-it Notes":                                "Miscellaneous",
    "Tape":                                         "Miscellaneous",
}

CATEGORY_ORDER = [
    "Cleaning Sprays",
    "Scrubbing & Sponges",
    "Wipes",
    "Soaps & Detergents",
    "Toilet & Bathroom",
    "Mopping & Sweeping",
    "Cloths & Towels",
    "Bags",
    "Gloves & PPE",
    "Air Care",
    "Miscellaneous",
]


def load_categories():
    if os.path.exists(CATEGORIES_FILE):
        with open(CATEGORIES_FILE) as f:
            return json.load(f)
    cats = dict(DEFAULT_CATEGORIES)
    save_categories(cats)
    return cats


def save_categories(cats):
    with open(CATEGORIES_FILE, "w") as f:
        json.dump(cats, f, indent=2)


def apply_categories(items):
    cats = load_categories()
    for item in items:
        item["section"] = cats.get(item["name"], "Miscellaneous")
    def _rank(item):
        try:
            return CATEGORY_ORDER.index(item["section"])
        except ValueError:
            return len(CATEGORY_ORDER)
    items.sort(key=_rank)
    return items


def get_items(cfg):
    return apply_categories(load_inventory(cfg))


SHARED_CSS = """
*{box-sizing:border-box}
body{font-family:-apple-system,sans-serif;margin:0;background:#f0f4f8}
.card{max-width:720px;margin:32px auto;background:white;
      border-radius:10px;box-shadow:0 2px 12px rgba(0,0,0,.12);overflow:hidden}
.hdr{background:#1f4e79;color:white;padding:20px 24px}
.hdr h1{margin:0;font-size:20px}
.hdr p{margin:4px 0 0;font-size:13px;opacity:.8}
.admin-bar{display:flex;align-items:center;gap:4px;padding:8px 14px;
           background:#16385a;flex-wrap:wrap}
.admin-bar a{font-size:13px;text-decoration:none;padding:5px 12px;
             border-radius:4px;color:#b8d4f0}
.admin-bar a:hover{background:rgba(255,255,255,.15);color:white}
.admin-bar .active{background:rgba(255,255,255,.22);color:white;font-weight:700}
.admin-bar .divider{color:#2d5070;font-size:18px;padding:0 2px;user-select:none}
.sub-bar{display:flex;gap:10px;padding:8px 24px;background:#eaf0f8;
         border-bottom:1px solid #d0dcea;font-size:13px;align-items:center}
.sub-bar a{color:#1f4e79;text-decoration:none;font-weight:600}
.sub-bar a:hover{text-decoration:underline}
.sub-bar .sep{color:#aaa}
.body{padding:20px 24px}
table{width:100%;border-collapse:collapse}
th{background:#1f4e79;color:white;padding:9px 8px;font-size:13px;font-weight:600}
.sec-row td{background:#dce6f1;color:#1f4e79;font-weight:700;font-size:13px;padding:7px 10px}
.btns{display:flex;gap:10px;margin-top:20px;flex-wrap:wrap}
.btn-blue{background:#1f4e79;color:white;border:none;padding:10px 22px;border-radius:6px;
          font-size:14px;font-weight:600;cursor:pointer;text-decoration:none;display:inline-block}
.btn-blue:hover{background:#16385a}
.btn-green{background:#1a7a3c;color:white;border:none;padding:10px 22px;border-radius:6px;
           font-size:14px;font-weight:600;cursor:pointer;text-decoration:none;display:inline-block}
.btn-green:hover{background:#155e2e}
.btn-grey{background:#e4e8ed;color:#333;border:none;padding:10px 18px;border-radius:6px;
          font-size:14px;cursor:pointer;text-decoration:none;display:inline-block}
.btn-grey:hover{background:#d0d5db}
.sec-title{font-size:15px;font-weight:700;color:#1f4e79;margin:24px 0 10px;
           border-bottom:2px solid #dce6f1;padding-bottom:6px}
input[type=text],input[type=number]{padding:6px 8px;border:1px solid #ccc;
  border-radius:4px;font-size:14px;width:100%}
label{font-size:13px;color:#555;display:block;margin-bottom:4px}
.new-grid{display:grid;grid-template-columns:1fr 100px 120px;gap:12px;
          align-items:end;margin-bottom:8px}
"""


# ── File version ───────────────────────────────────────────────────────────────

def get_inventory_version():
    mtime = 0
    seen = set()
    for cfg in INVENTORIES.values():
        f = cfg["file"]
        if f not in seen and os.path.exists(f):
            seen.add(f)
            try:
                mtime = max(mtime, int(os.path.getmtime(f)))
            except OSError:
                pass
    return mtime


# ── Quantity helpers ────────────────────────────────────────────────────────────

def parse_qty(raw):
    if raw is None:
        return 0, "", "none"
    if isinstance(raw, (int, float)):
        return int(raw), "", "int"
    s = str(raw).strip()
    parts = s.split(None, 1)
    try:
        return int(parts[0]), (parts[1].strip() if len(parts) > 1 else ""), "text"
    except (ValueError, IndexError):
        return 0, s, "none"


def format_back(qty, unit, fmt):
    if fmt == "text" and unit:
        return f"{qty} {unit}"
    return qty


def is_section_header(name):
    return bool(name) and ord(str(name)[0]) > 0x2000


# ── Excel helpers ───────────────────────────────────────────────────────────────

def build_debug_page():
    html = f"""<!DOCTYPE html><html><head>
  <title>Debug — Raw Excel Data</title><meta charset="utf-8">
  <style>{SHARED_CSS}.card{{max-width:1100px}}
  td,th{{padding:5px 8px;font-size:12px;border:1px solid #ddd}}
  td:first-child{{color:#888;width:40px}}</style>
</head><body><div class="card">
  <div class="hdr"><h1>Raw Excel Data</h1>
    <p>All columns, all rows, both read modes</p></div>
  {admin_bar(active_tab="debug")}
  <div class="body">"""

    for key, cfg in INVENTORIES.items():
        html += f'<div class="sec-title">{cfg["label"]} — file: {cfg["file"]} — sheet: "{cfg["sheet"]}"</div>'
        for mode_label, do_mode in [("data_only=True", True), ("data_only=False", False)]:
            html += f'<p style="font-size:13px;font-weight:600;margin:10px 0 4px">{mode_label}</p>'
            try:
                wb = openpyxl.load_workbook(cfg["file"], keep_vba=False, data_only=do_mode)
                html += f'<p style="font-size:12px;color:#666">Sheets in file: {wb.sheetnames}</p>'
                if cfg["sheet"] not in wb.sheetnames:
                    html += f'<p style="color:red">Sheet "{cfg["sheet"]}" NOT FOUND</p>'
                    continue
                ws = wb[cfg["sheet"]]
                html += f'<p style="font-size:12px;color:#666">max_row={ws.max_row} max_col={ws.max_column}</p>'
                html += '<table><tr><th>Row</th>'
                for c in range(1, (ws.max_column or 5) + 1):
                    from openpyxl.utils import get_column_letter
                    html += f'<th>Col {get_column_letter(c)}</th>'
                html += '</tr>'
                for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
                    html += f'<tr><td>{row_idx}</td>'
                    for cell in row:
                        val = "" if cell is None else str(cell)
                        color = "background:#ffffcc" if val else ""
                        html += f'<td style="{color}">{val}</td>'
                    html += '</tr>'
                html += '</table>'
            except Exception as e:
                html += f'<p style="color:red">Error: {e}</p>'

    html += '</div></div></body></html>'
    return html

def load_inventory(cfg):
    wb = openpyxl.load_workbook(cfg["file"], keep_vba=False, data_only=True)
    ws = wb[cfg["sheet"]]
    col = cfg["qty_col"]
    items = []
    current_section = ""
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name = row[0]
        if not name:
            continue
        name_str = str(name).strip()
        if is_section_header(name_str):
            current_section = name_str
            continue
        raw_qty = row[col - 1]
        qty, unit, fmt = parse_qty(raw_qty)
        if not unit and cfg.get("unit_col"):
            raw_unit = row[cfg["unit_col"] - 1]
            if raw_unit:
                unit = str(raw_unit).strip()
        items.append({
            "name":    name_str,
            "qty":     qty,
            "unit":    unit,
            "display": str(int(qty)) if isinstance(raw_qty, float) else (str(raw_qty) if raw_qty is not None else "0"),
            "row":     row_idx,
            "fmt":     fmt,
            "section": current_section,
        })
    return items


def save_checkout(cfg, items, selections):
    wb = openpyxl.load_workbook(cfg["file"], keep_vba=cfg["keep_vba"])
    ws = wb[cfg["sheet"]]
    col = cfg["qty_col"]
    for idx, take_qty in selections:
        item = items[idx]
        ws.cell(item["row"], col).value = format_back(
            item["qty"] - take_qty, item["unit"], item["fmt"])
    log_ws = wb["Log"] if "Log" in wb.sheetnames else wb.create_sheet("Log")
    if log_ws.cell(1, 1).value != "Timestamp":
        log_ws.insert_rows(1)
        for c, h in enumerate(["Timestamp", "Item Name", "Qty Taken", "Unit"], 1):
            log_ws.cell(1, c).value = h
    ts = datetime.now().strftime("%m/%d/%Y %I:%M %p")
    for idx, take_qty in selections:
        item = items[idx]
        log_ws.append([ts, item["name"], take_qty, item["unit"] or ""])
    wb.save(cfg["file"])


def save_restock(cfg, items, additions, new_item):
    wb = openpyxl.load_workbook(cfg["file"], keep_vba=cfg["keep_vba"])
    ws = wb[cfg["sheet"]]
    col = cfg["qty_col"]
    for idx, add_qty in additions:
        item = items[idx]
        ws.cell(item["row"], col).value = format_back(
            item["qty"] + add_qty, item["unit"], item["fmt"])
    if new_item:
        last = ws.max_row
        ws.cell(last + 1, 1).value = new_item["name"]
        ws.cell(last + 1, col).value = new_item["qty"]
        if new_item.get("unit") and cfg.get("unit_col"):
            ws.cell(last + 1, cfg["unit_col"]).value = new_item["unit"]
    wb.save(cfg["file"])


def save_delete_item(cfg, item):
    wb = openpyxl.load_workbook(cfg["file"], keep_vba=cfg["keep_vba"])
    ws = wb[cfg["sheet"]]
    ws.delete_rows(item["row"])
    wb.save(cfg["file"])


def save_edit_item(cfg, item, new_name, new_unit):
    wb = openpyxl.load_workbook(cfg["file"], keep_vba=cfg["keep_vba"])
    ws = wb[cfg["sheet"]]
    ws.cell(item["row"], 1).value = new_name
    if cfg.get("unit_col"):
        ws.cell(item["row"], cfg["unit_col"]).value = new_unit
    wb.save(cfg["file"])


# ── Shared fragments ────────────────────────────────────────────────────────────

def err_block(errors):
    if not errors:
        return ""
    li = "".join(f"<li>{e}</li>" for e in errors)
    return (f'<div style="background:#fff0f0;border:1px solid #e88;border-radius:6px;'
            f'padding:12px 16px;margin-bottom:16px;color:#900">'
            f'<strong>Please fix:</strong><ul style="margin:6px 0 0">{li}</ul></div>')


def checkout_rows(items, action_prefix):
    rows = ""
    last_sec = None
    for i, item in enumerate(items):
        if item["section"] != last_sec:
            last_sec = item["section"]
            rows += f'<tr class="sec-row"><td colspan="4">{last_sec}</td></tr>'
        bg = "#f5f8fd" if i % 2 == 0 else "#fff"
        low = " style='color:#cc0000;font-weight:bold'" if item["qty"] <= 2 else ""
        rows += f"""<tr style="background:{bg}" data-item="{item['name']}">
          <td style="padding:9px 8px;width:36px">
            <input type="checkbox" id="c{i}" name="check_{i}" onchange="tog({i})"
                   style="width:16px;height:16px;cursor:pointer">
          </td>
          <td style="padding:9px 8px">
            <label for="c{i}" style="cursor:pointer;font-size:14px">{item['name']}</label>
          </td>
          <td style="padding:9px 8px;text-align:center;font-size:14px"{low}>{item['display']}{(' ' + item['unit']) if item['unit'] and item['fmt'] != 'text' else ''}</td>
          <td style="padding:9px 8px;text-align:center">
            <input type="number" id="q{i}" name="qty_{i}" min="1" max="{item['qty']}"
                   disabled placeholder="—"
                   style="width:60px;text-align:center;padding:4px;font-size:14px;
                          border:1px solid #ccc;border-radius:4px;background:#ffffaa">
          </td>
        </tr>"""
    return rows


TOGGLE_JS = """<script>
function tog(i){var c=document.getElementById('c'+i),q=document.getElementById('q'+i);
  q.disabled=!c.checked; if(!c.checked)q.value=''; else{q.focus();q.select();}}
</script>"""

META_REFRESH = '<meta http-equiv="refresh" content="5">'

def ts():
    return datetime.now().strftime("%-I:%M:%S %p")

CANCEL_REFRESH_JS = """<script>
document.addEventListener('DOMContentLoaded',function(){
  function stopRefresh(){
    var m=document.querySelector('meta[http-equiv="refresh"]');
    if(m)m.parentNode.removeChild(m);
  }
  document.querySelectorAll('input,select,textarea').forEach(function(el){
    el.addEventListener('focus',stopRefresh);
  });
});
</script>"""

SEARCH_BAR = """<div style="margin-bottom:14px">
  <input type="text" id="srch" placeholder="&#128269; Search items..."
         style="width:100%;padding:9px 12px;font-size:14px;border:1px solid #ccc;
                border-radius:6px;box-sizing:border-box"
         oninput="srchFn(this.value)">
</div>
<script>
function srchFn(q){
  q=q.toLowerCase();
  document.querySelectorAll('tr[data-item]').forEach(function(r){
    r.style.display=(!q||r.dataset.item.toLowerCase().indexOf(q)>=0)?'':'none';
  });
  document.querySelectorAll('tr.sec-row').forEach(function(s){
    var sib=s.nextElementSibling,vis=false;
    while(sib&&!sib.classList.contains('sec-row')){
      if(sib.style.display!=='none')vis=true;
      sib=sib.nextElementSibling;
    }
    s.style.display=(vis||!q)?'':'none';
  });
}
</script>"""

SAVE_FORM_JS = """<script>
(function(){
  var fields=['new_name','new_qty','new_unit'];
  fields.forEach(function(n){
    var el=document.querySelector('[name="'+n+'"]');
    if(!el)return;
    var v=sessionStorage.getItem('mgr_'+n);
    if(v)el.value=v;
    el.addEventListener('input',function(){sessionStorage.setItem('mgr_'+n,el.value);});
  });
  var f=document.querySelector('form');
  if(f)f.addEventListener('submit',function(){
    fields.forEach(function(n){sessionStorage.removeItem('mgr_'+n);});
  });
})();
</script>"""


# ── Standalone pages (no admin access) ─────────────────────────────────────────

def build_home_page():
    def card(key, cfg):
        return f"""
        <a href="/{key}" style="text-decoration:none;flex:1;min-width:180px">
          <div style="background:white;border-radius:10px;
                      box-shadow:0 2px 10px rgba(0,0,0,.1);padding:28px 20px;
                      text-align:center;transition:box-shadow .2s;cursor:pointer;
                      border:2px solid transparent"
               onmouseover="this.style.boxShadow='0 4px 18px rgba(0,0,0,.18)'"
               onmouseout="this.style.boxShadow='0 2px 10px rgba(0,0,0,.1)'">
            <div style="font-size:40px;margin-bottom:10px">&#128722;</div>
            <div style="font-size:20px;font-weight:700;color:#1f4e79;margin-bottom:6px">
              {cfg['label']}</div>
            <div style="font-size:13px;color:#888">Tap to check out supplies</div>
          </div>
        </a>"""

    loc_cards = "".join(card(k, v) for k, v in INVENTORIES.items())

    return f"""<!DOCTYPE html><html><head>
  <title>Inventory — Select Location</title><meta charset="utf-8">
  {META_REFRESH}
  <style>{SHARED_CSS}.card{{max-width:760px}}</style>
</head><body>
  <div class="card">
    <div class="hdr">
      <h1>&#128722; Cleaning Supply Inventory</h1>
      <p>Select your location to check out supplies &nbsp;·&nbsp; Updated {ts()}</p>
    </div>
    <div class="body">
      <div style="display:flex;gap:16px;flex-wrap:wrap">
        {loc_cards}
      </div>
    </div>
  </div>
</body></html>"""


def standalone_subnav(key, active_tab):
    co = "<strong>Checkout</strong>" if active_tab == "checkout" \
         else f'<a href="/{key}">Checkout</a>'
    mg = "<strong>Manage Inventory</strong>" if active_tab == "manage" \
         else f'<a href="/{key}/manage">Manage Inventory</a>'
    return (f'<div class="sub-bar"><a href="/">&#8592; Home</a>'
            f'<span class="sep">|</span>{co}<span class="sep">|</span>{mg}</div>')


def build_standalone_checkout(key, items, errors=None):
    cfg = INVENTORIES[key]
    rows = checkout_rows(items, f"/{key}")
    return f"""<!DOCTYPE html><html><head>
  <title>Checkout — {cfg['label']}</title><meta charset="utf-8">
  {META_REFRESH}
  <style>{SHARED_CSS}</style>
</head><body>
  <div class="card">
    <div class="hdr">
      <h1>&#128722; {cfg['label']} — Checkout</h1>
      <p>Check items and enter the quantity you are taking &nbsp;·&nbsp; Updated {ts()}</p>
    </div>
    {standalone_subnav(key, "checkout")}
    <div class="body">
      {err_block(errors)}
      {SEARCH_BAR}
      <form method="POST" action="/{key}/checkout">
        <table>
          <tr>
            <th style="width:36px"></th>
            <th style="text-align:left">Item</th>
            <th style="width:130px">In Stock</th>
            <th style="width:100px">Qty to Take</th>
          </tr>
          {rows}
        </table>
        <div class="btns">
          <button class="btn-blue" type="submit">Process Checkout</button>
          <a href="/" class="btn-grey">Cancel</a>
        </div>
      </form>
    </div>
  </div>
  {TOGGLE_JS}{CANCEL_REFRESH_JS}
</body></html>"""


def build_standalone_manage(key, items, message=None, errors=None):
    cfg = INVENTORIES[key]
    cat_options = "".join(f'<option value="{c}">{c}</option>' for c in CATEGORY_ORDER)
    rows = ""
    last_sec = None
    for i, item in enumerate(items):
        if item["section"] != last_sec:
            last_sec = item["section"]
            rows += f'<tr class="sec-row"><td colspan="4">{last_sec}</td></tr>'
        bg = "#f5f8fd" if i % 2 == 0 else "#fff"
        name = item['name']
        rows += f"""<tr style="background:{bg}" data-item="{name}">
          <td style="padding:9px 8px;font-size:14px">{name}</td>
          <td style="padding:9px 8px;text-align:center;font-size:14px">{item['display']}{(' ' + item['unit']) if item['unit'] and item['fmt'] != 'text' else ''}</td>
          <td style="padding:9px 8px;text-align:center">
            <input type="number" name="add_{i}" min="1" placeholder="—"
                   style="width:65px;text-align:center;padding:4px;font-size:14px;
                          border:1px solid #ccc;border-radius:4px;background:#e8f5e9">
            {('<span style="font-size:12px;color:#888;margin-left:4px">' + item['unit'] + '</span>') if item['unit'] else ''}
          </td>
          <td style="padding:9px 8px;text-align:center">
            <a href="/{key}/edit-item?row={item['row']}"
               style="display:inline-block;background:#1f4e79;color:white;padding:4px 10px;
                      border-radius:4px;font-size:12px;text-decoration:none;margin-right:4px">Edit</a>
            <button type="button"
              onclick="if(confirm('Delete {name}? This cannot be undone.'))location.href='/{key}/delete-item?row={item['row']}'"
              style="background:#c0392b;color:white;border:none;padding:4px 10px;
                     border-radius:4px;font-size:12px;cursor:pointer">Delete</button>
          </td>
        </tr>"""

    msg = ""
    if message:
        msg = (f'<div style="background:#e8f5e9;border:1px solid #6abf7b;border-radius:6px;'
               f'padding:12px 16px;margin-bottom:16px;color:#1a5c2e;font-weight:600">'
               f'&#10003; {message}</div>')

    return f"""<!DOCTYPE html><html><head>
  <title>Manage — {cfg['label']}</title><meta charset="utf-8">
  <style>{SHARED_CSS}.new-grid{{grid-template-columns:1fr 90px 110px 140px}}</style>
</head><body>
  <div class="card">
    <div class="hdr"><h1>Manage Inventory — {cfg['label']}</h1>
      <p>Restock existing items or add new ones</p></div>
    {standalone_subnav(key, "manage")}
    <div class="body">
      {msg}{err_block(errors)}
      <form method="POST" action="/{key}/restock">
        <div class="sec-title">Restock Existing Items</div>
        <p style="font-size:13px;color:#666;margin:0 0 12px">
          Enter how many to <strong>add</strong>. Leave blank to skip.</p>
        {SEARCH_BAR}
        <table>
          <tr>
            <th style="text-align:left">Item</th>
            <th style="width:140px">Current Stock</th>
            <th style="width:110px">Add Qty</th>
            <th style="width:80px"></th>
          </tr>
          {rows}
        </table>
        <div class="sec-title">Add a New Item</div>
        <div class="new-grid">
          <div><label>Item Name</label>
            <input type="text" name="new_name" placeholder="e.g. Sponges"></div>
          <div><label>Quantity</label>
            <input type="number" name="new_qty" min="1" placeholder="0"></div>
          <div><label>Unit (optional)</label>
            <input type="text" name="new_unit" placeholder="e.g. box"></div>
          <div><label>Category</label>
            <select name="new_cat" style="width:100%;padding:6px 8px;border:1px solid #ccc;
                    border-radius:4px;font-size:14px">
              <option value="">— select —</option>
              {cat_options}
            </select></div>
        </div>
        <div class="btns">
          <button class="btn-green" type="submit">Save Changes</button>
          <a href="/{key}" class="btn-grey">Back to Checkout</a>
        </div>
      </form>
      <p style="font-size:11px;color:#bbb;margin-top:20px">
        Reading: {cfg['file']} &nbsp;·&nbsp; Sheet: {cfg['sheet']} &nbsp;·&nbsp;
        Modified: {datetime.fromtimestamp(os.path.getmtime(cfg['file'])).strftime('%-m/%-d/%Y %-I:%M:%S %p') if os.path.exists(cfg['file']) else 'not found'}
      </p>
    </div>
  </div>
  {SAVE_FORM_JS}
</body></html>"""


def build_standalone_success(key, items_taken):
    cfg = INVENTORIES[key]
    li = "".join(
        f"<li style='padding:4px 0'><strong>{n}</strong> — {q}{(' '+u) if u else ''}</li>"
        for n, q, u in items_taken)
    return f"""<!DOCTYPE html><html><head>
  <title>Done — {cfg['label']}</title><meta charset="utf-8">
  <style>{SHARED_CSS}.card{{max-width:480px;text-align:center}}
  ul{{text-align:left;display:inline-block}}h2{{color:#1f4e79;margin-top:0}}</style>
</head><body>
  <div class="card" style="margin:60px auto;padding:32px">
    <h2>&#10003; Checkout Complete</h2>
    <p style="color:#555">{cfg['label']} inventory updated.</p>
    <ul>{li}</ul><br>
    <div class="btns" style="justify-content:center">
      <a href="/{key}" class="btn-blue">New Checkout</a>
      <a href="/{key}/manage" class="btn-green">Manage Inventory</a>
      <a href="/" class="btn-grey">Home</a>
    </div>
  </div>
</body></html>"""


# ── Admin pages ─────────────────────────────────────────────────────────────────

def admin_bar(active_key=None, active_tab=None):
    links = '<a href="/" class="' + ("active" if not active_key and active_tab != "add-location" else "") + '">&#9776; Overview</a>'
    links += '<span class="divider">|</span>'
    for key, cfg in INVENTORIES.items():
        is_active = active_key == key and active_tab == "checkout"
        cls = "active" if is_active else ""
        links += f'<a href="/{key}" class="{cls}">{cfg["label"]}</a>'
    links += '<span class="divider">|</span>'
    for key, cfg in INVENTORIES.items():
        is_active = active_key == key and active_tab == "manage"
        cls = "active" if is_active else ""
        links += f'<a href="/{key}/manage" class="{cls}">Manage {cfg["label"]}</a>'
    links += '<span class="divider">|</span>'
    cls = "active" if active_tab == "add-location" else ""
    links += f'<a href="/add-location" class="{cls}">+ Add Location</a>'
    return f'<div class="admin-bar">{links}</div>'


def build_admin_overview(message=None, errors=None):
    all_sections = {}
    section_order = []
    low_stock = []  # (item_name, location_label, qty, unit)

    for key, cfg in INVENTORIES.items():
        try:
            items = get_items(cfg)
        except Exception:
            continue
        for item in items:
            sec = item["section"]
            if sec not in all_sections:
                all_sections[sec] = {}
                section_order.append(sec)
            if item["name"] not in all_sections[sec]:
                all_sections[sec][item["name"]] = {}
            all_sections[sec][item["name"]][key] = (item["display"], item["qty"], item["unit"], item["fmt"])
            if item["qty"] <= 2:
                unit_str = (' ' + item["unit"]) if item["unit"] else ''
                low_stock.append((item["name"], cfg["label"], item["qty"], unit_str))

    keys = list(INVENTORIES.keys())
    th = "".join(f'<th style="width:120px">{INVENTORIES[k]["label"]}</th>' for k in keys)

    rows = ""
    for sec in section_order:
        rows += f'<tr class="sec-row"><td colspan="{1+len(keys)}">{sec}</td></tr>'
        for i, (name, loc_data) in enumerate(all_sections[sec].items()):
            bg = "#f5f8fd" if i % 2 == 0 else "#fff"
            cells = ""
            for k in keys:
                if k in loc_data:
                    disp, qty, unit, fmt = loc_data[k]
                    unit_str = (' ' + unit) if unit and fmt != 'text' else ''
                    style = "color:#cc0000;font-weight:bold" if qty <= 2 else "color:#222"
                    cells += f'<td style="padding:8px;text-align:center;font-size:13px;{style}">{disp}{unit_str}</td>'
                else:
                    cells += '<td style="padding:8px;text-align:center;color:#ccc">—</td>'
            rows += (f'<tr style="background:{bg}">'
                     f'<td style="padding:8px 10px;font-size:13px">{name}</td>'
                     f'{cells}</tr>')

    # Build per-location low-stock banners
    alert_banner = ""
    for key, cfg in INVENTORIES.items():
        loc_low = [(name, qty, unit) for name, label, qty, unit in low_stock if label == cfg["label"]]
        if not loc_low:
            continue
        loc_low.sort(key=lambda x: x[1])
        alert_rows = "".join(
            f'<tr><td style="padding:5px 10px;font-size:13px;font-weight:600">{name}</td>'
            f'<td style="padding:5px 10px;font-size:13px;font-weight:700;color:#a00">{qty}{unit}</td></tr>'
            for name, qty, unit in loc_low
        )
        alert_banner += f"""
        <div style="background:#fff0f0;border:2px solid #e74c3c;border-radius:8px;
                    padding:14px 18px;margin-bottom:14px">
          <div style="display:flex;align-items:center;gap:8px;margin-bottom:10px">
            <span style="font-size:18px">&#9888;</span>
            <strong style="font-size:14px;color:#a00">{cfg["label"]} — {len(loc_low)} item(s) at 2 or below</strong>
          </div>
          <table style="width:100%;border-collapse:collapse">
            <tr>
              <th style="text-align:left;padding:4px 10px;font-size:12px;color:#888;
                         font-weight:600;background:none;border-bottom:1px solid #f5c0c0">Item</th>
              <th style="text-align:left;padding:4px 10px;font-size:12px;color:#888;
                         font-weight:600;background:none;border-bottom:1px solid #f5c0c0">Qty Left</th>
            </tr>
            {alert_rows}
          </table>
        </div>"""

    msg_block = ""
    if message:
        msg_block = (f'<div style="background:#e8f5e9;border:1px solid #6abf7b;border-radius:6px;'
                     f'padding:12px 16px;margin-bottom:16px;color:#1a5c2e;font-weight:600">'
                     f'&#10003; {message}</div>')

    loc_options = "".join(
        f'<option value="{k}">{cfg["label"]}</option>'
        for k, cfg in INVENTORIES.items())
    cat_options = "".join(f'<option value="{c}">{c}</option>' for c in CATEGORY_ORDER)

    add_item_form = f"""
    <div class="sec-title" style="margin-top:28px">Add Item to Location(s)</div>
    <div style="background:#f8fafc;border:1px solid #dce6f1;border-radius:8px;padding:16px 18px">
      <form method="POST" action="/add-item-to-locations">
        <div style="display:grid;grid-template-columns:1fr 80px 110px 140px 150px;
                    gap:10px;align-items:end;margin-bottom:12px">
          <div><label>Item Name</label>
            <input type="text" name="new_name" placeholder="e.g. Sponges"></div>
          <div><label>Quantity</label>
            <input type="number" name="new_qty" min="0" value="0" placeholder="0"></div>
          <div><label>Unit (optional)</label>
            <input type="text" name="new_unit" placeholder="e.g. boxes"></div>
          <div><label>Category</label>
            <select name="new_cat" style="width:100%;padding:6px 8px;border:1px solid #ccc;
                    border-radius:4px;font-size:14px">
              <option value="">— select —</option>
              {cat_options}
            </select></div>
          <div><label>Add to</label>
            <select name="target_loc" style="width:100%;padding:6px 8px;border:1px solid #ccc;
                    border-radius:4px;font-size:14px">
              <option value="__all__">All Locations</option>
              {loc_options}
            </select></div>
        </div>
        <button class="btn-green" type="submit" style="padding:8px 20px">Add Item</button>
      </form>
    </div>"""

    return f"""<!DOCTYPE html><html><head>
  <title>Admin — Overview</title><meta charset="utf-8">
  {META_REFRESH}
  <style>{SHARED_CSS}.card{{max-width:860px}}</style>
</head><body>
  <div class="card">
    <div class="hdr"><h1>&#9776; Admin — All Locations</h1>
      <p>Live stock levels across every location — red means 2 or fewer remaining &nbsp;·&nbsp; Updated {ts()}</p></div>
    {admin_bar()}
    <div class="body">
      {msg_block}{err_block(errors)}
      {alert_banner}
      <table>
        <tr><th style="text-align:left">Item</th>{th}</tr>
        {rows}
      </table>
      <p style="font-size:12px;color:#aaa;margin-top:10px">
        Updates automatically when stock changes.</p>
      {add_item_form}
    </div>
  </div>
</body></html>"""


def build_admin_checkout(key, items, errors=None):
    cfg = INVENTORIES[key]
    rows = checkout_rows(items, f"/{key}")
    return f"""<!DOCTYPE html><html><head>
  <title>Admin Checkout — {cfg['label']}</title><meta charset="utf-8">
  {META_REFRESH}
  <style>{SHARED_CSS}</style>
</head><body>
  <div class="card">
    <div class="hdr"><h1>&#128722; {cfg['label']} — Checkout</h1>
      <p>Admin view — check items and enter the quantity being taken &nbsp;·&nbsp; Updated {ts()}</p></div>
    {admin_bar(key, "checkout")}
    <div class="body">
      {err_block(errors)}
      {SEARCH_BAR}
      <form method="POST" action="/{key}/checkout">
        <table>
          <tr>
            <th style="width:36px"></th>
            <th style="text-align:left">Item</th>
            <th style="width:130px">In Stock</th>
            <th style="width:100px">Qty to Take</th>
          </tr>
          {rows}
        </table>
        <div class="btns">
          <button class="btn-blue" type="submit">Process Checkout</button>
        </div>
      </form>
    </div>
  </div>
  {TOGGLE_JS}{CANCEL_REFRESH_JS}
</body></html>"""


def build_admin_manage(key, items, message=None, errors=None):
    cfg = INVENTORIES[key]
    cat_options = "".join(f'<option value="{c}">{c}</option>' for c in CATEGORY_ORDER)
    rows = ""
    last_sec = None
    for i, item in enumerate(items):
        if item["section"] != last_sec:
            last_sec = item["section"]
            rows += f'<tr class="sec-row"><td colspan="4">{last_sec}</td></tr>'
        bg = "#f5f8fd" if i % 2 == 0 else "#fff"
        name = item['name']
        rows += f"""<tr style="background:{bg}" data-item="{name}">
          <td style="padding:9px 8px;font-size:14px">{name}</td>
          <td style="padding:9px 8px;text-align:center;font-size:14px">{item['display']}{(' ' + item['unit']) if item['unit'] and item['fmt'] != 'text' else ''}</td>
          <td style="padding:9px 8px;text-align:center">
            <input type="number" name="add_{i}" min="1" placeholder="—"
                   style="width:65px;text-align:center;padding:4px;font-size:14px;
                          border:1px solid #ccc;border-radius:4px;background:#e8f5e9">
            {('<span style="font-size:12px;color:#888;margin-left:4px">' + item['unit'] + '</span>') if item['unit'] else ''}
          </td>
          <td style="padding:9px 8px;text-align:center">
            <a href="/{key}/edit-item?row={item['row']}"
               style="display:inline-block;background:#1f4e79;color:white;padding:4px 10px;
                      border-radius:4px;font-size:12px;text-decoration:none;margin-right:4px">Edit</a>
            <button type="button"
              onclick="if(confirm('Delete {name}? This cannot be undone.'))location.href='/{key}/delete-item?row={item['row']}'"
              style="background:#c0392b;color:white;border:none;padding:4px 10px;
                     border-radius:4px;font-size:12px;cursor:pointer">Delete</button>
          </td>
        </tr>"""

    msg = ""
    if message:
        msg = (f'<div style="background:#e8f5e9;border:1px solid #6abf7b;border-radius:6px;'
               f'padding:12px 16px;margin-bottom:16px;color:#1a5c2e;font-weight:600">'
               f'&#10003; {message}</div>')

    return f"""<!DOCTYPE html><html><head>
  <title>Manage — {cfg['label']}</title><meta charset="utf-8">
  <style>{SHARED_CSS}.new-grid{{grid-template-columns:1fr 90px 110px 140px}}</style>
</head><body>
  <div class="card">
    <div class="hdr"><h1>Manage Inventory — {cfg['label']}</h1>
      <p>Restock existing items or add new ones</p></div>
    {admin_bar(key, "manage")}
    <div class="body">
      {msg}{err_block(errors)}
      <form method="POST" action="/{key}/restock">
        <div class="sec-title">Restock Existing Items</div>
        <p style="font-size:13px;color:#666;margin:0 0 12px">
          Enter how many to <strong>add</strong>. Leave blank to skip.</p>
        {SEARCH_BAR}
        <table>
          <tr>
            <th style="text-align:left">Item</th>
            <th style="width:140px">Current Stock</th>
            <th style="width:110px">Add Qty</th>
            <th style="width:80px"></th>
          </tr>
          {rows}
        </table>
        <div class="sec-title">Add a New Item</div>
        <div class="new-grid">
          <div><label>Item Name</label>
            <input type="text" name="new_name" placeholder="e.g. Sponges"></div>
          <div><label>Quantity</label>
            <input type="number" name="new_qty" min="1" placeholder="0"></div>
          <div><label>Unit (optional)</label>
            <input type="text" name="new_unit" placeholder="e.g. box"></div>
          <div><label>Category</label>
            <select name="new_cat" style="width:100%;padding:6px 8px;border:1px solid #ccc;
                    border-radius:4px;font-size:14px">
              <option value="">— select —</option>
              {cat_options}
            </select></div>
        </div>
        <div class="btns">
          <button class="btn-green" type="submit">Save Changes</button>
        </div>
      </form>
      <p style="font-size:11px;color:#bbb;margin-top:20px">
        Reading: {cfg['file']} &nbsp;·&nbsp; Sheet: {cfg['sheet']} &nbsp;·&nbsp;
        Modified: {datetime.fromtimestamp(os.path.getmtime(cfg['file'])).strftime('%-m/%-d/%Y %-I:%M:%S %p') if os.path.exists(cfg['file']) else 'not found'}
      </p>
    </div>
  </div>
  {SAVE_FORM_JS}
</body></html>"""


def build_admin_success(key, items_taken):
    cfg = INVENTORIES[key]
    li = "".join(
        f"<li style='padding:4px 0'><strong>{n}</strong> — {q}{(' '+u) if u else ''}</li>"
        for n, q, u in items_taken)
    return f"""<!DOCTYPE html><html><head>
  <title>Done — {cfg['label']}</title><meta charset="utf-8">
  <style>{SHARED_CSS}.card{{max-width:520px;text-align:center}}
  ul{{text-align:left;display:inline-block}}h2{{color:#1f4e79;margin-top:0}}</style>
</head><body>
  <div class="card" style="margin:60px auto;overflow:visible">
    <div class="hdr"><h1>&#10003; Checkout Complete — {cfg['label']}</h1></div>
    {admin_bar(key, "checkout")}
    <div style="padding:24px;text-align:center">
      <ul>{li}</ul><br>
      <div class="btns" style="justify-content:center">
        <a href="/{key}" class="btn-blue">New Checkout</a>
        <a href="/{key}/manage" class="btn-green">Manage {cfg['label']}</a>
        <a href="/" class="btn-grey">Overview</a>
      </div>
    </div>
  </div>
</body></html>"""


def build_admin_add_location(message=None, errors=None):
    item_list = "".join(
        f"<li style='padding:2px 0;font-size:13px'>{name} — {qty} {unit}</li>"
        for name, qty, unit in get_template_items())

    loc_rows = ""
    for key, cfg in INVENTORIES.items():
        lbl = cfg['label']
        loc_rows += f"""<tr>
          <td style="padding:10px 8px;font-size:14px;font-weight:600">{lbl}</td>
          <td style="padding:10px 8px;font-size:13px;color:#666">/{key}</td>
          <td style="padding:10px 8px;text-align:right">
            <form method="POST" action="/delete-location" style="display:inline"
                  onsubmit="return confirm('Delete {lbl}? This cannot be undone.')">
              <input type="hidden" name="key" value="{key}">
              <button type="submit"
                style="background:#c0392b;color:white;border:none;padding:5px 12px;
                       border-radius:4px;font-size:13px;cursor:pointer">
                Delete
              </button>
            </form>
          </td>
        </tr>"""

    msg = ""
    if message:
        msg = (f'<div style="background:#e8f5e9;border:1px solid #6abf7b;border-radius:6px;'
               f'padding:12px 16px;margin-bottom:16px;color:#1a5c2e;font-weight:600">'
               f'&#10003; {message}</div>')

    return f"""<!DOCTYPE html><html><head>
  <title>Admin — Manage Locations</title><meta charset="utf-8">
  <style>{SHARED_CSS}.card{{max-width:680px}}</style>
</head><body>
  <div class="card">
    <div class="hdr"><h1>Manage Locations</h1>
      <p>Add or remove locations</p></div>
    {admin_bar(active_tab="add-location")}
    <div class="body">
      {msg}{err_block(errors)}
      <div class="sec-title">Existing Locations</div>
      <table>
        <tr>
          <th style="text-align:left">Name</th>
          <th style="text-align:left">Staff Link</th>
          <th style="width:100px"></th>
        </tr>
        {loc_rows}
      </table>
      <div class="sec-title">Add New Location</div>
      <form method="POST" action="/add-location">
        <div style="max-width:300px;margin-bottom:12px">
          <label>Location Name</label>
          <input type="text" name="label" placeholder="e.g. Downtown" autofocus>
        </div>
        <p style="font-size:13px;color:#666;margin:0 0 8px">
          New locations start with these standard quantities (adjust via Manage Inventory):</p>
        <ul style="margin:0 0 16px;padding-left:20px;columns:2;column-gap:40px">
          {item_list}
        </ul>
        <div class="btns">
          <button class="btn-green" type="submit">Create Location</button>
          <a href="/" class="btn-grey">Cancel</a>
        </div>
      </form>
    </div>
  </div>
</body></html>"""


def build_edit_page(key, item, admin, errors=None):
    cfg = INVENTORIES[key]
    cats = load_categories()
    current_cat = cats.get(item["name"], "Miscellaneous")
    cat_options = "".join(
        f'<option value="{c}" {"selected" if c == current_cat else ""}>{c}</option>'
        for c in CATEGORY_ORDER
    )
    nav = admin_bar(key, "manage") if admin else standalone_subnav(key, "manage")
    unit_val = item["unit"] if item["unit"] and item["fmt"] != "text" else ""
    return f"""<!DOCTYPE html><html><head>
  <title>Edit Item — {cfg['label']}</title><meta charset="utf-8">
  <style>{SHARED_CSS}.card{{max-width:520px}}</style>
</head><body>
  <div class="card">
    <div class="hdr"><h1>Edit Item — {cfg['label']}</h1>
      <p>Update the name, unit, or category for this item</p></div>
    {nav}
    <div class="body">
      {err_block(errors)}
      <form method="POST" action="/{key}/edit-item">
        <input type="hidden" name="row" value="{item['row']}">
        <input type="hidden" name="old_name" value="{item['name']}">
        <div style="margin-bottom:14px">
          <label style="font-size:13px;color:#555;display:block;margin-bottom:4px">Item Name</label>
          <input type="text" name="new_name" value="{item['name']}"
                 style="width:100%;padding:8px;border:1px solid #ccc;border-radius:4px;font-size:14px">
        </div>
        <div style="margin-bottom:14px">
          <label style="font-size:13px;color:#555;display:block;margin-bottom:4px">Unit (optional)</label>
          <input type="text" name="new_unit" value="{unit_val}"
                 style="width:100%;padding:8px;border:1px solid #ccc;border-radius:4px;font-size:14px"
                 placeholder="e.g. bottles">
        </div>
        <div style="margin-bottom:20px">
          <label style="font-size:13px;color:#555;display:block;margin-bottom:4px">Category</label>
          <select name="new_cat" style="width:100%;padding:8px;border:1px solid #ccc;
                  border-radius:4px;font-size:14px">
            {cat_options}
          </select>
        </div>
        <div class="btns">
          <button class="btn-green" type="submit">Save Changes</button>
          <a href="/{key}/manage" class="btn-grey">Cancel</a>
        </div>
      </form>
    </div>
  </div>
</body></html>"""


# ── Server ──────────────────────────────────────────────────────────────────────

shutdown_event = threading.Event()


class BaseHandler(http.server.BaseHTTPRequestHandler):

    def _checkout(self, key, form, admin):
        cfg = INVENTORIES[key]
        items = get_items(cfg)
        errors, selections, summary = [], [], []

        for i, item in enumerate(items):
            if f"check_{i}" not in form:
                continue
            qty_str = form.get(f"qty_{i}", [""])[0].strip()
            if not qty_str:
                errors.append(f"{item['name']}: enter a quantity")
                continue
            if not qty_str.isdigit() or int(qty_str) <= 0:
                errors.append(f"{item['name']}: must be a whole number greater than 0")
                continue
            qty = int(qty_str)
            if qty > item["qty"]:
                errors.append(f"{item['name']}: only {item['display']} in stock")
                continue
            selections.append((i, qty))
            summary.append((item["name"], qty, item["unit"]))

        if not selections and not errors:
            errors.append("Please check at least one item.")

        build_page = build_admin_checkout if admin else build_standalone_checkout
        if errors:
            self._ok(build_page(key, items, errors))
            return
        try:
            save_checkout(cfg, items, selections)
            if admin:
                self._ok(build_admin_success(key, summary))
            else:
                self._ok(build_standalone_success(key, summary))
        except PermissionError:
            self._ok(build_page(key, items, [
                f"{os.path.basename(cfg['file'])} is open in Excel. Close it and try again."]))
        except Exception as e:
            self._ok(build_page(key, items, [f"Error saving: {e}"]))

    def _restock(self, key, form, admin=True):
        cfg = INVENTORIES[key]
        items = get_items(cfg)
        errors, additions, new_item = [], [], None

        for i, item in enumerate(items):
            val = form.get(f"add_{i}", [""])[0].strip()
            if not val:
                continue
            if not val.isdigit() or int(val) <= 0:
                errors.append(f"{item['name']}: must be a whole number greater than 0")
                continue
            additions.append((i, int(val)))

        new_name = form.get("new_name", [""])[0].strip()
        new_qty  = form.get("new_qty",  [""])[0].strip()
        new_unit = form.get("new_unit", [""])[0].strip()
        new_cat  = form.get("new_cat",  [""])[0].strip() or "Miscellaneous"

        if new_name or new_qty:
            if not new_name:
                errors.append("New item: enter an item name")
            if not new_qty or not new_qty.isdigit() or int(new_qty) <= 0:
                errors.append("New item: enter a valid starting quantity")
            if not errors:
                new_item = {"name": new_name, "qty": int(new_qty), "unit": new_unit}

        build_manage = build_admin_manage if admin else build_standalone_manage

        if not additions and not new_item and not errors:
            self._ok(build_manage(key, items, errors=[
                "Nothing to save — enter a restock quantity or fill in a new item."]))
            return
        if errors:
            self._ok(build_manage(key, items, errors=errors))
            return
        try:
            save_restock(cfg, items, additions, new_item)
            if new_item:
                cats = load_categories()
                cats[new_item["name"]] = new_cat
                save_categories(cats)
            parts = []
            if additions:
                parts.append(f"{len(additions)} item(s) restocked")
            if new_item:
                parts.append(f"'{new_item['name']}' added")
            self._ok(build_manage(key, get_items(cfg),
                     message=" and ".join(parts) + "."))
        except PermissionError:
            self._ok(build_manage(key, items, errors=[
                f"{os.path.basename(cfg['file'])} is open in Excel. Close it and try again."]))
        except Exception as e:
            self._ok(build_manage(key, items, errors=[f"Error saving: {e}"]))

    def _edit_item_get(self, key, admin):
        from urllib.parse import urlparse, parse_qs
        params = parse_qs(urlparse(self.path).query)
        try:
            row = int(params.get("row", [0])[0])
        except (ValueError, IndexError):
            row = 0
        cfg = INVENTORIES[key]
        items = load_inventory(cfg)
        item = next((it for it in items if it["row"] == row), None)
        if not item:
            build_manage = build_admin_manage if admin else build_standalone_manage
            self._ok(build_manage(key, get_items(cfg), errors=["Item not found."]))
            return
        self._ok(build_edit_page(key, item, admin))

    def _edit_item_post(self, key, form, admin):
        cfg = INVENTORIES[key]
        build_manage = build_admin_manage if admin else build_standalone_manage
        try:
            row = int(form.get("row", ["0"])[0])
        except ValueError:
            self._ok(build_manage(key, get_items(cfg), errors=["Invalid row."]))
            return
        old_name = form.get("old_name", [""])[0].strip()
        new_name = form.get("new_name", [""])[0].strip()
        new_unit = form.get("new_unit", [""])[0].strip()
        new_cat  = form.get("new_cat",  ["Miscellaneous"])[0].strip() or "Miscellaneous"
        if not new_name:
            items = load_inventory(cfg)
            item = next((it for it in items if it["row"] == row), None)
            self._ok(build_edit_page(key, item, admin, errors=["Item name cannot be blank."]))
            return
        items = load_inventory(cfg)
        item = next((it for it in items if it["row"] == row), None)
        if not item:
            self._ok(build_manage(key, get_items(cfg), errors=["Item not found."]))
            return
        try:
            save_edit_item(cfg, item, new_name, new_unit)
            cats = load_categories()
            if old_name in cats:
                del cats[old_name]
            cats[new_name] = new_cat
            save_categories(cats)
            self._ok(build_manage(key, get_items(cfg),
                     message=f"'{new_name}' updated successfully."))
        except PermissionError:
            self._ok(build_edit_page(key, item, admin, errors=[
                f"{os.path.basename(cfg['file'])} is open in Excel. Close it and try again."]))
        except Exception as e:
            self._ok(build_edit_page(key, item, admin, errors=[f"Error saving: {e}"]))

    def _delete_item(self, key, admin):
        from urllib.parse import urlparse, parse_qs
        params = parse_qs(urlparse(self.path).query)
        try:
            row = int(params.get("row", [0])[0])
        except (ValueError, IndexError):
            row = 0
        cfg = INVENTORIES[key]
        items = load_inventory(cfg)
        item = next((it for it in items if it["row"] == row), None)
        build_manage = build_admin_manage if admin else build_standalone_manage
        if not item:
            self._ok(build_manage(key, items, errors=["Item not found."]))
            return
        try:
            save_delete_item(cfg, item)
            self._ok(build_manage(key, get_items(cfg),
                     message=f"'{item['name']}' has been deleted."))
        except Exception as e:
            self._ok(build_manage(key, items, errors=[f"Error deleting item: {e}"]))

    def _serve_version(self):
        data = json.dumps({"v": get_inventory_version()}).encode()
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", len(data))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(data)

    def _ok(self, html, code=200):
        data = html.encode()
        self.send_response(code)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", len(data))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, *args):
        pass


class UserHandler(BaseHandler):
    """Staff-facing server (PORT 8765). No admin routes."""

    def do_GET(self):
        p = urllib.parse.urlparse(self.path).path.rstrip("/") or "/"

        if p == "/api/version":
            self._serve_version()
        elif p == "/":
            self._ok(build_home_page())
        elif p in ("/quit", "/cancel"):
            self._ok("<html><body><script>window.close()</script>"
                     "<p>You may close this tab.</p></body></html>")
            threading.Timer(0.5, shutdown_event.set).start()
        elif p.lstrip("/") in INVENTORIES:
            key = p.lstrip("/")
            self._ok(build_standalone_checkout(key, get_items(INVENTORIES[key])))
        elif p.count("/") == 2 and p.split("/")[1] in INVENTORIES and p.endswith("/manage"):
            key = p.split("/")[1]
            self._ok(build_standalone_manage(key, get_items(INVENTORIES[key])))
        elif p.count("/") == 2 and p.split("/")[1] in INVENTORIES and p.endswith("/delete-item"):
            key = p.split("/")[1]
            self._delete_item(key, admin=False)
        elif p.count("/") == 2 and p.split("/")[1] in INVENTORIES and p.endswith("/edit-item"):
            key = p.split("/")[1]
            self._edit_item_get(key, admin=False)
        else:
            self._ok("<h1>Not found</h1>", 404)

    def do_POST(self):
        length = int(self.headers.get("Content-Length", 0))
        form = urllib.parse.parse_qs(self.rfile.read(length).decode())
        parts = self.path.strip("/").split("/")

        if len(parts) == 2 and parts[0] in INVENTORIES and parts[1] == "checkout":
            self._checkout(parts[0], form, admin=False)
        elif len(parts) == 2 and parts[0] in INVENTORIES and parts[1] == "restock":
            self._restock(parts[0], form, admin=False)
        elif len(parts) == 2 and parts[0] in INVENTORIES and parts[1] == "edit-item":
            self._edit_item_post(parts[0], form, admin=False)
        else:
            self._ok("<h1>Not found</h1>", 404)


class AdminHandler(BaseHandler):
    """Admin-only server (ADMIN_PORT 8766). Full access to all locations."""

    def do_GET(self):
        p = urllib.parse.urlparse(self.path).path.rstrip("/") or "/"

        if p == "/api/version":
            self._serve_version()
        elif p == "/debug":
            self._ok(build_debug_page())
        elif p == "/":
            self._ok(build_admin_overview())
        elif p == "/add-location":
            self._ok(build_admin_add_location())
        elif p.lstrip("/") in INVENTORIES:
            key = p.lstrip("/")
            self._ok(build_admin_checkout(key, get_items(INVENTORIES[key])))
        elif p.count("/") == 2 and p.split("/")[1] in INVENTORIES and p.endswith("/manage"):
            key = p.split("/")[1]
            self._ok(build_admin_manage(key, get_items(INVENTORIES[key])))
        elif p.count("/") == 2 and p.split("/")[1] in INVENTORIES and p.endswith("/delete-item"):
            key = p.split("/")[1]
            self._delete_item(key, admin=True)
        elif p.count("/") == 2 and p.split("/")[1] in INVENTORIES and p.endswith("/edit-item"):
            key = p.split("/")[1]
            self._edit_item_get(key, admin=True)
        else:
            self._ok("<h1>Not found</h1>", 404)

    def do_POST(self):
        length = int(self.headers.get("Content-Length", 0))
        form = urllib.parse.parse_qs(self.rfile.read(length).decode())
        parts = self.path.strip("/").split("/")

        if self.path.strip("/") == "add-item-to-locations":
            self._add_item_to_locations(form)
        elif self.path.strip("/") == "add-location":
            self._add_location(form)
        elif self.path.strip("/") == "delete-location":
            self._delete_location(form)
        elif len(parts) == 2 and parts[0] in INVENTORIES and parts[1] == "checkout":
            self._checkout(parts[0], form, admin=True)
        elif len(parts) == 2 and parts[0] in INVENTORIES and parts[1] == "restock":
            self._restock(parts[0], form, admin=True)
        elif len(parts) == 2 and parts[0] in INVENTORIES and parts[1] == "edit-item":
            self._edit_item_post(parts[0], form, admin=True)
        else:
            self._ok("<h1>Not found</h1>", 404)

    def _add_item_to_locations(self, form):
        new_name   = form.get("new_name",   [""])[0].strip()
        new_qty    = form.get("new_qty",    ["0"])[0].strip()
        new_unit   = form.get("new_unit",   [""])[0].strip()
        new_cat    = form.get("new_cat",    ["Miscellaneous"])[0].strip() or "Miscellaneous"
        target_loc = form.get("target_loc", ["__all__"])[0].strip()

        errors = []
        if not new_name:
            errors.append("Enter an item name.")
        qty = int(new_qty) if new_qty.isdigit() else 0

        if errors:
            self._ok(build_admin_overview(errors=errors))
            return

        target_keys = list(INVENTORIES.keys()) if target_loc == "__all__" \
                      else ([target_loc] if target_loc in INVENTORIES else [])
        if not target_keys:
            self._ok(build_admin_overview(errors=["Invalid location selected."]))
            return

        new_item = {"name": new_name, "qty": qty, "unit": new_unit}
        added_to, skipped = [], []

        for key in target_keys:
            cfg = INVENTORIES[key]
            items = load_inventory(cfg)
            if any(it["name"].lower() == new_name.lower() for it in items):
                skipped.append(cfg["label"])
                continue
            try:
                save_restock(cfg, items, [], new_item)
                added_to.append(cfg["label"])
            except Exception as e:
                errors.append(f"{cfg['label']}: {e}")

        if added_to:
            cats = load_categories()
            cats[new_name] = new_cat
            save_categories(cats)

        parts = []
        if added_to:
            parts.append(f"'{new_name}' added to {', '.join(added_to)}")
        if skipped:
            parts.append(f"already existed in {', '.join(skipped)} (skipped)")
        message = ". ".join(parts) + "." if parts else None
        self._ok(build_admin_overview(message=message, errors=errors or None))

    def _add_location(self, form):
        label = form.get("label", [""])[0].strip()
        if not label:
            self._ok(build_admin_add_location(errors=["Please enter a location name."]))
            return

        key = re.sub(r"[^a-z0-9]+", "-", label.lower()).strip("-")
        if not key:
            self._ok(build_admin_add_location(errors=["Invalid location name."]))
            return
        if key in INVENTORIES:
            self._ok(build_admin_add_location(
                errors=[f"A location with the key '{key}' already exists."]))
            return

        try:
            create_location_sheet(label)
        except Exception as e:
            self._ok(build_admin_add_location(errors=[f"Could not create sheet: {e}"]))
            return

        INVENTORIES[key] = {
            "label":    label,
            "file":     COMBINED_FILE,
            "sheet":    label,
            "qty_col":  2,
            "unit_col": 3,
            "keep_vba": False,
        }
        save_locations()
        self._ok(build_admin_add_location(
            message=f"'{label}' created successfully. Staff link: /{key}"))

    def _delete_location(self, form):
        key = form.get("key", [""])[0].strip()
        if key not in INVENTORIES:
            self._ok(build_admin_add_location(errors=["Location not found."]))
            return
        cfg = INVENTORIES[key]
        label = cfg["label"]
        # Remove the sheet from Inventory.xlsx if that's where it lives
        if cfg["file"] == COMBINED_FILE:
            try:
                wb = openpyxl.load_workbook(COMBINED_FILE)
                if cfg["sheet"] in wb.sheetnames:
                    del wb[cfg["sheet"]]
                    wb.save(COMBINED_FILE)
            except Exception:
                pass
        del INVENTORIES[key]
        save_locations()
        self._ok(build_admin_add_location(message=f"'{label}' has been removed."))


def get_local_ip():
    import subprocess
    # Try Tailscale first (utun interfaces), then fall back to LAN
    for iface in ("utun0", "utun1", "utun2", "utun3", "en0", "en1", "en2"):
        try:
            result = subprocess.run(
                ["ipconfig", "getifaddr", iface],
                capture_output=True, text=True, timeout=2)
            ip = result.stdout.strip()
            if ip and ip.startswith("100."):
                return ip
        except Exception:
            continue
    for iface in ("en0", "en1", "en2"):
        try:
            result = subprocess.run(
                ["ipconfig", "getifaddr", iface],
                capture_output=True, text=True, timeout=2)
            ip = result.stdout.strip()
            if ip:
                return ip
        except Exception:
            continue
    return "localhost"


def run_server(handler_class, port):
    try:
        server = http.server.HTTPServer(("0.0.0.0", port), handler_class)
        print(f"  Server started on port {port} ✓")
        server.serve_forever()
    except OSError as e:
        print(f"  ERROR on port {port}: {e}")
        print(f"  Try running: lsof -i :{port}  to see what is using that port.")


def main():
    ip = get_local_ip()
    staff_url = f"http://localhost:{PORT}"
    print("\n" + "="*52)
    print("  STAFF LINKS (share each link with that location):")
    for key, cfg in INVENTORIES.items():
        print(f"  {cfg['label']:10}  http://{ip}:{PORT}/{key}")
    print()
    print("  ADMIN LINK (keep this private):")
    print(f"  http://{ip}:{ADMIN_PORT}")
    print("="*52)
    migrate_to_combined()
    migrate_separate_locations()
    print("  Starting servers...")
    threading.Thread(target=run_server, args=(AdminHandler, ADMIN_PORT), daemon=True).start()
    threading.Timer(0.8, lambda: webbrowser.open(staff_url)).start()
    try:
        run_server(UserHandler, PORT)
    except KeyboardInterrupt:
        pass
    print("App closed.")


if __name__ == "__main__":
    main()
